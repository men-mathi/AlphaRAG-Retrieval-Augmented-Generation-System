import os
import zipfile
import csv
from reportlab.pdfgen import canvas
from openpyxl import Workbook

# Ensure docs directory exists
os.makedirs('docs', exist_ok=True)

# 1. Create PDF
pdf_path = 'docs/test_pdf.pdf'
c = canvas.Canvas(pdf_path)
c.drawString(100, 750, "This is a test PDF document.")
c.drawString(100, 730, "The secret agent code is PDF-AGENT-007.")
c.save()

# 2. Create Excel (XLSX)
xlsx_path = 'docs/test_excel.xlsx'
wb = Workbook()
ws = wb.active
ws['A1'] = "Test Sheet"
ws['A2'] = "This is a test Excel document."
ws['A3'] = "The secret location is Sector 7G."
wb.save(xlsx_path)

# 3. Create DOCX (Word)
docx_path = 'docs/test_word.docx'
with zipfile.ZipFile(docx_path, 'w') as f:
    xml_content = """<?xml version="1.0" encoding="utf-8" standalone="yes"?>
    <w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
        <w:p>
            <w:r>
                <w:t>This is a test Word document. The project code name is AlphaOmega99.</w:t>
            </w:r>
        </w:p>
    </w:document>"""
    f.writestr('word/document.xml', xml_content)

# 4. Create PPTX (PowerPoint)
pptx_path = 'docs/test_powerpoint.pptx'
with zipfile.ZipFile(pptx_path, 'w') as f:
    xml_content = """<?xml version="1.0" encoding="utf-8" standalone="yes"?>
    <sld xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
        <a:t>This is a test PowerPoint slide. The lead presenter is Bob Jones.</a:t>
    </sld>"""
    f.writestr('ppt/slides/slide1.xml', xml_content)

# 5. Create CSV
csv_path = 'docs/test_csv.csv'
with open(csv_path, 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerow(['Name', 'Role', 'AccessCode'])
    writer.writerow(['Charlie Brown', 'AI Specialist', 'AccessCode CB123'])

# 6. Create HTML
html_path = 'docs/test_html.html'
with open(html_path, 'w', encoding='utf-8') as f:
    f.write("""<html>
    <head><title>Test Page</title></head>
    <body>
        <h1>HTML Test Content</h1>
        <p>This is extracted from an HTML page. The contact person is Diana Prince.</p>
    </body>
    </html>""")

print("Successfully created mock test documents inside docs/ directory.")
